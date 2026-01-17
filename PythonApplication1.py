# -*- coding: utf-8 -*-


from pathlib import Path
import math
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from bvpnewton import solve_bvp_newton, build_linear_guess

# -------------------------
# Исходные данные варианта 35
# -------------------------
class Params:
    W = 3435.0          # м/с
    P = 9520.0          # Н (9.52 кН)
    h_isl = 153e3       # м
    R = 1738e3          # м
    mu = 4.903e12       # м^3/с^2 (важно: именно 4.903e12, как в эталоне)

    tv = 14.0           # с
    dt_base = 0.1       # базовый шаг
    dt_min = 1e-6       # минимальный шаг уточнения
    tol_v = 1e-8        # точность по целевой скорости

# Параметры управления (из эталона / результата оптимизации)
T1 = 323.17
T2 = 661.2
THETA_END1_DEG = -0.5     # theta_s(t1), град
THETA2_DEG = -16.443888720254627       # theta_s на 2-м АУТ, град

# -------------------------
# Форматирование Excel
# -------------------------
NUM_FMT = "0.000000000000000"  # 15 знаков после запятой
INT_FMT = "0"                  # для "№ п/п", чтобы не было ######

thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="F2F2F2")
title_font = Font(bold=True, size=12)
header_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

def _cell_display_str(cell):
    """Оценка длины того, что Excel реально покажет, с учётом number_format."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, (int, float, np.floating, np.integer)):
        fmt = cell.number_format or ""
        if fmt == NUM_FMT:
            try:
                return f"{float(v):.15f}"
            except Exception:
                return str(v)
        if fmt == INT_FMT:
            try:
                return str(int(round(float(v))))
            except Exception:
                return str(v)
        return str(v)
    return str(v)

def autosize(ws, min_width=10, max_width=60):
    """Подгон ширины колонок так, чтобы не появлялись ###### из-за формата 15 знаков."""
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            s = _cell_display_str(ws.cell(row=row, column=col))
            if s:
                max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(col)].width = min(max(min_width, max_len + 2), max_width)

def write_title(ws, r, title, ncols):
    cell = ws.cell(row=r, column=1, value=title)
    cell.font = title_font
    cell.alignment = left
    if ncols > 1:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    return r + 1

def write_header(ws, r, headers):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
    return r + 1

def write_values(ws, r, headers, values):
    for c, (h, v) in enumerate(zip(headers, values), start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = border
        cell.alignment = center
        if isinstance(v, (int, float, np.floating, np.integer)):
            cell.number_format = INT_FMT if h.strip() == "№ п/п" else NUM_FMT
    return r + 1

# -------------------------
# Численная модель (RK4)
# -------------------------
class Phase:
    SEG1 = "SEG1"
    COAST = "COAST"
    SEG2 = "SEG2"

def rk4_step(t, state, dt, phase, p: Params, dtheta1, theta1_end, theta2):
    Beta = p.P / p.W

    def theta_s(tloc):
        if phase == Phase.SEG1:
            if tloc <= p.tv:
                return math.pi / 2
            return math.pi / 2 + dtheta1 * (tloc - p.tv)
        if phase == Phase.COAST:
            return theta1_end
        return theta2

    thrust = phase in (Phase.SEG1, Phase.SEG2)

    def deriv(tloc, s):
        vx, vy, x, y, m = s
        rx, ry = x, p.R + y
        r = math.hypot(rx, ry)

        ax_g = -p.mu * rx / (r**3)
        ay_g = -p.mu * ry / (r**3)

        if thrust:
            ts = theta_s(tloc)
            ax_t = (p.P / m) * math.cos(ts)
            ay_t = (p.P / m) * math.sin(ts)
            dm = -Beta
        else:
            ax_t = ay_t = 0.0
            dm = 0.0

        return np.array([ax_g + ax_t, ay_g + ay_t, vx, vy, dm], dtype=float)

    k1 = deriv(t, state)
    k2 = deriv(t + dt/2, state + dt*k1/2)
    k3 = deriv(t + dt/2, state + dt*k2/2)
    k4 = deriv(t + dt,   state + dt*k3)
    return state + (dt/6) * (k1 + 2*k2 + 2*k3 + k4)

def simulate_all(p: Params, t1=T1, t2=T2, theta_end1_deg=THETA_END1_DEG, theta2_deg=THETA2_DEG):
    theta1_end = math.radians(theta_end1_deg)
    theta2 = math.radians(theta2_deg)
    dtheta1 = theta1_end
    #(theta1_end - math.pi/2) / (t1 - p.tv)
    Vtarget = math.sqrt(p.mu / (p.R + p.h_isl))
    print (theta1_end)
    print (theta2)
    t = 0.0
    state = np.array([0.0, 0.0, 0.0, 0.0, 3000.0], dtype=float)  # vx, vy, x, y, m
    phase = Phase.SEG1
    dt_ref = p.dt_base

    duplicated_tv = False
    rows = []

    def theta_for_output(tcur, phasecur):
        # важно: в момент t2 (тяга ещё 0) угол уже "переставлен" в theta2 (как в эталоне)
        if phasecur == Phase.COAST and abs(tcur - t2) < 1e-12:
            return theta2
        if phasecur == Phase.SEG1:
            if tcur <= p.tv:
                return math.pi/2
            return math.pi/2 + dtheta1 * (tcur - p.tv)
        if phasecur == Phase.COAST:
            return theta1_end
        return theta2

    def thrust_for_output(phasecur):
        return p.P if phasecur in (Phase.SEG1, Phase.SEG2) else 0.0

    def record(tcur, st, phasecur):
        nonlocal duplicated_tv
        vx, vy, x, y, m = st
        ts = theta_for_output(tcur, phasecur)
        Pc = thrust_for_output(phasecur)
        rows.append((tcur, m, vx, vy, x, y, ts, Pc))
        # в эталонной "таблице расчёта" строка t=14 дублируется
        if (not duplicated_tv) and abs(tcur - p.tv) < 1e-12:
            rows.append((tcur, m, vx, vy, x, y, ts, Pc))
            duplicated_tv = True

    record(t, state, phase)

    def is_on_grid(tcur):
        return abs((tcur / p.dt_base) - round(tcur / p.dt_base)) < 1e-10

    def next_grid_time(tcur):
        k = math.floor(tcur / p.dt_base + 1e-12) + 1
        return k * p.dt_base

    while True:
        V = math.hypot(state[0], state[1])

        if abs(V - Vtarget) <= p.tol_v:
            break

        dt = dt_ref

        # попадание в t1/t2 ровно
        if phase == Phase.SEG1 and t + dt > t1:
            dt = t1 - t
        elif phase == Phase.COAST and t + dt > t2:
            dt = t2 - t

        # выравнивание обратно на сетку 0.1
        if dt_ref == p.dt_base:
            if abs(t - t1) < 1e-12 and not is_on_grid(t):
                dt = next_grid_time(t) - t
            if abs(t - t2) < 1e-12 and not is_on_grid(t):
                dt = next_grid_time(t) - t

        new_state = rk4_step(t, state, dt, phase, p, dtheta1, theta1_end, theta2)
        t_new = t + dt
        V_new = math.hypot(new_state[0], new_state[1])

        # уточнение шага десятками, если "перепрыгнули" целевую скорость
        if dt_ref > p.dt_min and V_new >= Vtarget:
            dt_ref = max(dt_ref / 10.0, p.dt_min)
            continue

        if dt_ref <= p.dt_min and V_new >= Vtarget:
            break

        t, state = t_new, new_state
        record(t, state, phase)

        # переключение фаз (важно: строка t2 должна быть с P=0)
        if phase == Phase.SEG1 and abs(t - t1) < 1e-12:
            phase = Phase.COAST
        elif phase == Phase.COAST and abs(t - t2) < 1e-12:
            phase = Phase.SEG2

    return pd.DataFrame(rows, columns=["t, с","m, кг","V_x, м/с","V_y, м/с","x, м","y, м","theta_s_rad","P, Н"])

def add_derived(df_raw: pd.DataFrame, p: Params):
    x = df_raw["x, м"].to_numpy()
    y = df_raw["y, м"].to_numpy()
    vx = df_raw["V_x, м/с"].to_numpy()
    vy = df_raw["V_y, м/с"].to_numpy()

    rx, ry = x, p.R + y
    r = np.sqrt(rx*rx + ry*ry)
    V = np.sqrt(vx*vx + vy*vy)

    phi = np.degrees(np.arctan2(rx, ry))
    Theta_s = np.degrees(np.arctan2(vy, vx))
    Theta_s = np.where(V < 1e-12, 90.0, Theta_s)

    theta_s = np.degrees(df_raw["theta_s_rad"].to_numpy())
    Theta = Theta_s + phi
    theta = theta_s + phi
    alpha = theta_s - Theta_s
    g = p.mu / (r*r)

    df = pd.DataFrame({
        "t, с": df_raw["t, с"],
        "m, кг": df_raw["m, кг"],
        "V_x, м/с": df_raw["V_x, м/с"],
        "V_y, м/с": df_raw["V_y, м/с"],
        "x, км": x/1000.0,
        "y, км": y/1000.0,
        "h, км": (r - p.R)/1000.0,
        "V, м/с": V,
        "r, км": r/1000.0,
        "theta_s, град": theta_s,
        "Theta_s, град": Theta_s,
        "alpha, град": alpha,
        "phi, град": phi,
        "theta, град": theta,
        "Theta, град": Theta,
        "g, м/с2": g,
        "P, Н": df_raw["P, Н"],
    })
    df.insert(0, "№ п/п", np.arange(0, len(df), dtype=int))
    return df

def trapz(t, f):
    return float(np.sum(0.5*(f[:-1] + f[1:]) * np.diff(t)))

def compute_summary_blocks(all_tbl: pd.DataFrame, p: Params, t1=T1, t2=T2):
    t = all_tbl["t, с"].to_numpy()
    m = all_tbl["m, кг"].to_numpy()

    m0 = 3000.0
    m1 = float(all_tbl.loc[np.isclose(all_tbl["t, с"], t1), "m, кг"].iloc[0])
    m2 = float(all_tbl.loc[np.isclose(all_tbl["t, с"], t2), "m, кг"].iloc[0])
    mf = float(m[-1])

    Pcol = all_tbl["P, Н"].to_numpy()
    alpha = np.deg2rad(all_tbl["alpha, град"].to_numpy())
    term_ctrl = (Pcol/m) * (1.0 - np.cos(alpha))
    term_g = all_tbl["g, м/с2"].to_numpy() * np.sin(np.deg2rad(all_tbl["Theta, град"].to_numpy()))

    mask1 = t <= t1 + 1e-12
    maskp = (t >= t1 - 1e-12) & (t <= t2 + 1e-12)
    mask2 = t >= t2 - 1e-12

    dVupr1 = trapz(t[mask1], term_ctrl[mask1])
    dVg1   = trapz(t[mask1], term_g[mask1])
    dVg_put = trapz(t[maskp], term_g[maskp])
    dVupr2 = trapz(t[mask2], term_ctrl[mask2])
    dVg2   = trapz(t[mask2], term_g[mask2])

    Vc1 = p.W * math.log(m0/m1)
    Vk1 = Vc1 - dVg1 - dVupr1
    Vk_put = Vk1 - dVg_put
    Vc2 = Vk_put + p.W * math.log(m2/mf)
    Vk2 = Vc2 - dVg2 - dVupr2

    return [
        ("Исходные данные", ["W_ist, м/с","P, кН","h_isl, км"], [p.W, p.P/1000.0, p.h_isl/1000.0]),
        ("Скорость Циолковского в конце 1-го АУТ", ["V_ц_к_1, м/с"], [Vc1]),
        ("Гравитационные потери и потери на управление на 1-м АУТ", ["dV_g_1, м/с","dV_упр_1, м/с"], [dVg1, dVupr1]),
        ("Характеристическая скорость в конце 1-го АУТ", ["V_к_1аут, м/с"], [Vk1]),
        ("Гравитационные потери на ПУТ", ["dV_g_пут, м/с"], [dVg_put]),
        ("Характеристическая скорость в конце ПУТ", ["V_к_пут, м/с"], [Vk_put]),
        ("Скорость Циолковского в конце 2-го АУТ", ["V_ц_к_2, м/с"], [Vc2]),
        ("Гравитационные потери и потери на управление на 2-м АУТ", ["dV_g_2, м/с","dV_упр_2, м/с"], [dVg2, dVupr2]),
        ("Характеристическая скорость в конце 2-го АУТ", ["V_к_2аут, м/с"], [Vk2]),
    ]

def orbit_elements_from_row(row, p: Params):
    x_m = float(row["x, км"]) * 1000.0
    y_m = float(row["y, км"]) * 1000.0
    vx = float(row["V_x, м/с"])
    vy = float(row["V_y, м/с"])

    rvec = np.array([x_m, p.R + y_m], dtype=float)
    vvec = np.array([vx, vy], dtype=float)

    r = np.linalg.norm(rvec)
    v = np.linalg.norm(vvec)

    h = np.cross(np.append(rvec, 0.0), np.append(vvec, 0.0))[2]
    eps = v*v/2.0 - p.mu/r
    a = -p.mu/(2.0*eps)
    e = math.sqrt(max(0.0, 1.0 - h*h/(p.mu*a)))

    rp = a*(1.0 - e)
    ra = a*(1.0 + e)
    return rp/1000.0, ra/1000.0, a/1000.0, e

def build_reduced(all_tbl: pd.DataFrame, t1=T1, t2=T2):
    # убрать дубликат t=14 для отбора строк
    uniq = all_tbl.drop_duplicates(subset=["t, с"], keep="first").reset_index(drop=True)
    tk = float(uniq["t, с"].iloc[-1])

    times = [0.0, 10.0, 14.0]
    tcur = 20.0
    while tcur <= math.floor(tk/10.0)*10.0 + 1e-9:
        times.append(tcur)
        tcur += 10.0
    for sp in [t1, t2, tk]:
        if all(abs(sp - tt) > 1e-6 for tt in times):
            times.append(sp)
    times = sorted(times)

    rows = []
    for tt in times:
        idx = int(np.argmin(np.abs(uniq["t, с"].to_numpy() - tt)))
        rows.append(uniq.iloc[idx].copy())

    red = pd.DataFrame(rows).reset_index(drop=True)
    red["№ п/п"] = np.arange(0, len(red), dtype=int)

    return red[[
        "№ п/п","t, с","m, кг","V_x, м/с","V_y, м/с","x, км","y, км","h, км",
        "V, м/с","r, км","theta_s, град","Theta_s, град","alpha, град","phi, град"
    ]]

def save_excels(all_tbl: pd.DataFrame, reduced_tbl: pd.DataFrame, orbit_tbl: pd.DataFrame, summary_blocks, out_traj1: Path, out_all: Path):
    # -------- Variant_35_traj_1.xlsx (в одной вкладке, как CSV-эталон) --------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Variant_35_traj_1"
    r = 1

    for title, header, values in summary_blocks:
        r = write_title(ws, r, title, ncols=max(3, len(header)))
        r = write_header(ws, r, header)
        r = write_values(ws, r, header, values)
        r += 1

    r = write_title(ws, r, "Элементы орбиты в конце активных участков", ncols=len(orbit_tbl.columns))
    orbit_headers = list(orbit_tbl.columns)
    r = write_header(ws, r, orbit_headers)
    for _, row in orbit_tbl.iterrows():
        r = write_values(ws, r, orbit_headers, row.tolist())
    r += 1

    r = write_title(ws, r, "Результаты расчета траектории выведения", ncols=len(reduced_tbl.columns))
    red_headers = list(reduced_tbl.columns)
    r = write_header(ws, r, red_headers)
    for _, row in reduced_tbl.iterrows():
        r = write_values(ws, r, red_headers, row.tolist())

    autosize(ws)
    wb.save(out_traj1)

    # -------- Variant_35_traj_1_All_Trajectory.xlsx --------
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "All_Trajectory"
    all_headers = list(all_tbl.columns)
    rr = write_header(ws2, 1, all_headers)
    for _, row in all_tbl.iterrows():
        rr = write_values(ws2, rr, all_headers, row.tolist())
    autosize(ws2)
    wb2.save(out_all)

def main():
    p = Params()

    bvp_result = solve_bvp_example()
    if not bvp_result.converged:
        raise RuntimeError(
            "BVP не сошлась: iterations={}, residual={:.3e}".format(
                bvp_result.iterations, bvp_result.residual_norm
            )
        )

    df_raw = simulate_all(p)
    all_tbl = add_derived(df_raw, p)

    summary_blocks = compute_summary_blocks(all_tbl, p)
    reduced_tbl = build_reduced(all_tbl)

    # элементы орбиты в конце 1-го и 2-го АУТ
    uniq = all_tbl.drop_duplicates(subset=["t, с"], keep="first").reset_index(drop=True)
    row1 = uniq.loc[np.isclose(uniq["t, с"], T1)].iloc[0]
    row2 = uniq.iloc[-1]

    rp1, ra1, a1, e1 = orbit_elements_from_row(row1, p)
    rp2, ra2, a2, e2 = orbit_elements_from_row(row2, p)

    orbit_tbl = pd.DataFrame(
        [
            [1, rp1, ra1, a1, e1],
            [2, rp2, ra2, a2, e2],
        ],
        columns=["№ АУТ","r_p, км","r_a, км","a, км","e"]
    )

    base = Path(__file__).resolve().parent
    save_excels(
        all_tbl=all_tbl,
        reduced_tbl=reduced_tbl,
        orbit_tbl=orbit_tbl,
        summary_blocks=summary_blocks,
        out_traj1=base / "Variant_35_traj_1.xlsx",
        out_all=base / "Variant_35_traj_1_All_Trajectory.xlsx",
    )
    print("OK: созданы Variant_35_traj_1.xlsx и Variant_35_traj_1_All_Trajectory.xlsx")

def solve_bvp_example():
    """Пример решения краевой задачи методом Ньютона (использует bvpnewton.py)."""
    def f(x, y):
        return np.array([y[1], -y[0]])

    def bc(y_a, y_b):
        return np.array([y_a[0], y_b[0] - 1.0])

    _x_guess, y_guess = build_linear_guess(a=0.0, b=1.0, y_a=[0.0, 0.0], y_b=[1.0, 0.0], n=50)
    result = solve_bvp_newton(f, bc, 0.0, 1.0, y_guess, n=50)
    print(
        "BVP: converged={}, iterations={}, residual={:.3e}".format(
            result.converged, result.iterations, result.residual_norm
        )
    )
    return result

if __name__ == "__main__":
    main()
